import json
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent
INPUT_FILE = BASE_DIR / "exercise_database/raw_excel/Übungen_Demoprogramm.xlsx"
OUT_FILE = BASE_DIR / "exercise_database/json/exercises_demo_structured.json"


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def row_values(ws, row_idx):
    return [clean(ws.cell(row=row_idx, column=c).value) for c in range(1, ws.max_column + 1)]


def parse_demo():
    wb = load_workbook(INPUT_FILE, data_only=True)
    ws = wb["Tabelle1"]

    exercises = []
    current = None

    for r in range(4, ws.max_row + 1):
        row = row_values(ws, r)

        krankheitsbild = row[0]
        stage = row[1]
        exercise_number = row[2]
        exercise_name = row[3]

        # Neue Übung beginnt, wenn Übungsnummer und Name vorhanden sind
        if exercise_number and exercise_name:
            if current:
                exercises.append(current)

            current = {
                "source": "Übungen_Demoprogramm.xlsx",
                "condition": krankheitsbild,
                "stage": stage,
                "exercise_number": exercise_number,
                "exercise_name": exercise_name,
                "execution_view": row[5],
                "equipment": row[6],
                "voiceover_description": row[7],
                "demo_repetitions": row[8],
                "training_repetitions": row[9],
                "quick_info": row[10],
                "pain_rules": [],
                "regression": {
                    "default": row[13],
                },
                "capture_problem": row[15],
                "software_problem": row[16],
                "criteria": [],
                "abort_criteria": [],
                "original_rows": [r],
            }

            # Basis-Schmerzregel aus Startzeile
            if row[11] or row[12]:
                current["pain_rules"].append({
                    "pain": row[11],
                    "action": row[12],
                    "voiceover": row[13],
                    "source_row": r
                })

            # Kriterien der Startzeile
            add_criteria_from_row(current, row, r)
            continue

        if current is None:
            continue

        current["original_rows"].append(r)

        # Schmerzregeln stehen in Spalten L-N = index 11-13
        if row[11] or row[12] or row[13]:
            current["pain_rules"].append({
                "pain": row[11],
                "action": row[12],
                "voiceover": row[13],
                "source_row": r
            })

        # Kriterien / Feedback / Abbruch ergänzen
        add_criteria_from_row(current, row, r)

    if current:
        exercises.append(current)

    return exercises


def add_criteria_from_row(current, row, source_row):
    """
    Aus der Demo-Tabelle:
    Spalten S-AB enthalten Kriterium/Feedbackblöcke.
    Spalten AD-AE enthalten Abbruchkriterien.
    """

    # Kriterienblöcke grob:
    # S = Kriterium-Label
    # U/V = Feedback 1
    # W/X = Feedback 2
    # Y/Z = Feedback 3
    criterion_label = row[18]

    feedback_pairs = [
        (row[20], row[21]),
        (row[22], row[23]),
        (row[24], row[25]),
        (row[26], row[27]),
    ]

    for condition, feedback in feedback_pairs:
        if condition or feedback:
            current["criteria"].append({
                "criterion_group": criterion_label,
                "condition": condition,
                "feedback": feedback,
                "source_row": source_row,
                "status": "original_table",
                "needs_mapping_to_mediapipe": True
            })

    abort_pairs = [
        row[29],
        row[30],
    ]

    for abort in abort_pairs:
        if abort:
            current["abort_criteria"].append({
                "condition": abort,
                "source_row": source_row,
                "status": "original_table",
                "needs_mapping_to_mediapipe": True
            })


if __name__ == "__main__":
    exercises = parse_demo()

    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(exercises, f, ensure_ascii=False, indent=2)

    print(f"Saved {len(exercises)} exercises to:")
    print(OUT_FILE)

    print("\nPreview:")
    for ex in exercises[:3]:
        print("-", ex["condition"], "|", ex["exercise_number"], "|", ex["exercise_name"])
        print("  criteria:", len(ex["criteria"]))
        print("  pain_rules:", len(ex["pain_rules"]))
        print("  abort:", len(ex["abort_criteria"]))