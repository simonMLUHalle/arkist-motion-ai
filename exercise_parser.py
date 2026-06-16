import json
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent

FILES = {
    "demo": BASE_DIR / "exercise_database/raw_excel/Übungen_Demoprogramm.xlsx",
    "overview": BASE_DIR / "exercise_database/raw_excel/Übungen_Gesamtübersicht.xlsx",
}

OUT_DIR = BASE_DIR / "exercise_database/json"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def cell_color(cell):
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.type == "rgb":
        return fill.fgColor.rgb
    return None


def parse_workbook(path):
    wb = load_workbook(path, data_only=True)
    result = {
        "file": str(path.name),
        "sheets": {}
    }

    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                row_data.append({
                    "cell": cell.coordinate,
                    "value": cell.value,
                    "fill_color": cell_color(cell)
                })
            rows.append(row_data)

        result["sheets"][ws.title] = {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_cells": [str(r) for r in ws.merged_cells.ranges],
            "rows": rows
        }

    return result


for name, path in FILES.items():
    print(f"Parsing {path}")

    parsed = parse_workbook(path)

    out_file = OUT_DIR / f"{name}_raw.json"

    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(parsed, f, ensure_ascii=False, indent=2)

    print(f"Saved: {out_file}")

print("Done.")